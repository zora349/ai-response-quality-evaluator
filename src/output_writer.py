import re

import pandas as pd

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from config import Path

def generate_excel_analysis(result_df: pd.DataFrame, output_path: Path):
    """
    生成多Sheet Excel分析报告：
    1. 汇总
    2. 错误类型统计
    3. 评分维度统计
    4. 低分样本
    5. 明细数据

    同时生成柱状图/饼图，并做基础样式美化。
    """

    # =========================
    # 1. 统计基础数据
    # =========================

    total = len(result_df)
    passed = len(result_df[result_df["auto_result"] == "合格"])
    failed = len(result_df[result_df["auto_result"] == "不合格"])
    pass_rate = passed / total if total else 0

    score_columns = [
        "auto_instruction_score",
        "auto_completeness_score",
        "auto_format_score",
        "auto_naturalness_score",
        "auto_accuracy_score",
        "auto_total_score",
    ]

    score_name_map = {
        "auto_instruction_score": "指令遵循",
        "auto_completeness_score": "回答完整性",
        "auto_format_score": "格式规范",
        "auto_naturalness_score": "表达自然度",
        "auto_accuracy_score": "准确性",
        "auto_total_score": "综合得分",
    }

    # =========================
    # 2. 汇总 Sheet 数据
    # =========================

    summary_df = pd.DataFrame({
        "指标": [
            "总样本数",
            "合格数量",
            "不合格数量",
            "合格率",
            "综合平均分",
            "指令遵循平均分",
            "回答完整性平均分",
            "格式规范平均分",
            "表达自然度平均分",
            "准确性平均分",
        ],
        "数值": [
            total,
            passed,
            failed,
            pass_rate,
            round(result_df["auto_total_score"].mean(), 2),
            round(result_df["auto_instruction_score"].mean(), 2),
            round(result_df["auto_completeness_score"].mean(), 2),
            round(result_df["auto_format_score"].mean(), 2),
            round(result_df["auto_naturalness_score"].mean(), 2),
            round(result_df["auto_accuracy_score"].mean(), 2),
        ]
    })

    result_distribution_df = pd.DataFrame({
        "结果": ["合格", "不合格"],
        "数量": [passed, failed],
    })

    # =========================
    # 3. 错误类型统计
    # =========================

    error_stats_df = (
        result_df["auto_error_type"]
        .value_counts()
        .reset_index()
    )
    error_stats_df.columns = ["错误类型", "数量"]
    error_stats_df["占比"] = error_stats_df["数量"] / total

    # =========================
    # 4. 评分维度统计
    # =========================

    score_stats_list = []

    for col in score_columns:
        score_stats_list.append({
            "评分维度": score_name_map[col],
            "平均分": round(result_df[col].mean(), 2),
            "最低分": result_df[col].min(),
            "最高分": result_df[col].max(),
        })

    score_stats_df = pd.DataFrame(score_stats_list)

    # =========================
    # 5. 低分样本
    # =========================

    low_score_df = result_df[
        (result_df["auto_result"] == "不合格") |
        (result_df["auto_total_score"] < 4)
    ].copy()

    low_score_columns = [
        "id",
        "user_input",
        "model_output",
        "parsed_rules",
        "auto_instruction_score",
        "auto_completeness_score",
        "auto_format_score",
        "auto_naturalness_score",
        "auto_accuracy_score",
        "auto_total_score",
        "auto_result",
        "auto_error_type",
        "auto_reason",
    ]

    low_score_columns = [col for col in low_score_columns if col in low_score_df.columns]
    low_score_df = low_score_df[low_score_columns].sort_values(
        by="auto_total_score",
        ascending=True
    )

    # =========================
    # 6. 写入 Excel 多个 Sheet
    # =========================

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总", index=False, startrow=0)
        result_distribution_df.to_excel(writer, sheet_name="汇总", index=False, startrow=13)

        error_stats_df.to_excel(writer, sheet_name="错误类型统计", index=False)
        score_stats_df.to_excel(writer, sheet_name="评分维度统计", index=False)
        low_score_df.to_excel(writer, sheet_name="低分样本", index=False)
        result_df.to_excel(writer, sheet_name="明细数据", index=False)

        wb = writer.book


        # =========================
        # 新增：图表分析 Sheet
        # =========================

        ws_chart = wb.create_sheet("图表分析")

        ws_chart["A1"] = "AI回复质量自动评估图表分析"
        ws_chart["A1"].font = Font(bold=True, size=16)

        # 1. 合格/不合格分布数据
        ws_chart["A3"] = "评估结果分布"
        ws_chart["A3"].font = Font(bold=True, size=12)

        ws_chart["A4"] = "结果"
        ws_chart["B4"] = "数量"
        ws_chart["A5"] = "合格"
        ws_chart["B5"] = passed
        ws_chart["A6"] = "不合格"
        ws_chart["B6"] = failed

        # 2. 错误类型统计数据
        ws_chart["D3"] = "错误类型分布"
        ws_chart["D3"].font = Font(bold=True, size=12)

        ws_chart["D4"] = "错误类型"
        ws_chart["E4"] = "数量"

        for idx, row in error_stats_df.iterrows():
            excel_row = idx + 5
            ws_chart[f"D{excel_row}"] = row["错误类型"]
            ws_chart[f"E{excel_row}"] = row["数量"]

        # 3. 评分维度统计数据
        ws_chart["A18"] = "评分维度平均分"
        ws_chart["A18"].font = Font(bold=True, size=12)

        ws_chart["A19"] = "评分维度"
        ws_chart["B19"] = "平均分"

        for idx, row in score_stats_df.iterrows():
            excel_row = idx + 20
            ws_chart[f"A{excel_row}"] = row["评分维度"]
            ws_chart[f"B{excel_row}"] = row["平均分"]

        # 4. 低分样本错误类型统计
        ws_chart["D18"] = "低分样本错误类型"
        ws_chart["D18"].font = Font(bold=True, size=12)

        low_error_stats_df = (
            low_score_df["auto_error_type"]
            .value_counts()
            .reset_index()
        )

        low_error_stats_df.columns = ["错误类型", "数量"]

        ws_chart["D19"] = "错误类型"
        ws_chart["E19"] = "数量"

        for idx, row in low_error_stats_df.iterrows():
            excel_row = idx + 20
            ws_chart[f"D{excel_row}"] = row["错误类型"]
            ws_chart[f"E{excel_row}"] = row["数量"]

        # =========================
        # 7. 通用样式函数
        # =========================

        def style_worksheet(ws):
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            header_font = Font(bold=True, color="000000")
            thin_side = Side(style="thin", color="D9D9D9")
            border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )

            ws.freeze_panes = "A2"

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=True
                    )
                    cell.border = border

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # 自动列宽，但限制最大宽度，避免太夸张
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_length = 0

                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max(max_length + 2, 10), 45)
                ws.column_dimensions[col_letter].width = adjusted_width

            # 行高
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 22

        for sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name])

        # =========================
        # 8. 汇总 Sheet 样式 + 图表
        # =========================

        ws_summary = wb["汇总"]

        ws_summary["D1"] = "合格/不合格分布"
        ws_summary["D1"].font = Font(bold=True, size=12)

        bar_chart = BarChart()
        bar_chart.title = "合格 / 不合格数量"
        bar_chart.y_axis.title = "数量"
        bar_chart.x_axis.title = "结果"

        data = Reference(ws_summary, min_col=2, min_row=14, max_row=16)
        cats = Reference(ws_summary, min_col=1, min_row=15, max_row=16)

        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.height = 7
        bar_chart.width = 12

        ws_summary.add_chart(bar_chart, "D2")

        pie_chart = PieChart()
        pie_chart.title = "合格率分布"

        pie_data = Reference(ws_summary, min_col=2, min_row=14, max_row=16)
        pie_cats = Reference(ws_summary, min_col=1, min_row=15, max_row=16)

        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_cats)
        pie_chart.height = 7
        pie_chart.width = 10

        ws_summary.add_chart(pie_chart, "D18")

        # 合格率格式
        ws_summary["B5"].number_format = "0.00%"

        # =========================
        # 9. 错误类型统计 Sheet 图表
        # =========================

        ws_error = wb["错误类型统计"]

        if ws_error.max_row >= 2:
            error_chart = BarChart()
            error_chart.title = "错误类型分布"
            error_chart.y_axis.title = "数量"
            error_chart.x_axis.title = "错误类型"

            error_data = Reference(ws_error, min_col=2, min_row=1, max_row=ws_error.max_row)
            error_cats = Reference(ws_error, min_col=1, min_row=2, max_row=ws_error.max_row)

            error_chart.add_data(error_data, titles_from_data=True)
            error_chart.set_categories(error_cats)
            error_chart.height = 8
            error_chart.width = 14

            ws_error.add_chart(error_chart, "E2")

            # 占比列格式
            for row in range(2, ws_error.max_row + 1):
                ws_error[f"C{row}"].number_format = "0.00%"

        # =========================
        # 10. 评分维度统计 Sheet 图表
        # =========================

        ws_score = wb["评分维度统计"]

        if ws_score.max_row >= 2:
            score_chart = BarChart()
            score_chart.title = "五项评分维度平均分"
            score_chart.y_axis.title = "平均分"
            score_chart.x_axis.title = "评分维度"

            score_data = Reference(ws_score, min_col=2, min_row=1, max_row=ws_score.max_row)
            score_cats = Reference(ws_score, min_col=1, min_row=2, max_row=ws_score.max_row)

            score_chart.add_data(score_data, titles_from_data=True)
            score_chart.set_categories(score_cats)
            score_chart.height = 8
            score_chart.width = 14

            ws_score.add_chart(score_chart, "F2")

        # =========================
        # 11. 明细数据 Sheet 条件格式
        # =========================

        ws_detail = wb["明细数据"]

        # 找到评分列位置
        detail_headers = [cell.value for cell in ws_detail[1]]

        for score_col in score_columns:
            if score_col in detail_headers:
                col_idx = detail_headers.index(score_col) + 1
                col_letter = get_column_letter(col_idx)

                ws_detail.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws_detail.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=1,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=3,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=5,
                        end_color="63BE7B",
                    )
                )

        # =========================
        # 12. 低分样本 Sheet 突出低分
        # =========================

        ws_low = wb["低分样本"]

        if ws_low.max_row >= 2:
            low_headers = [cell.value for cell in ws_low[1]]

            if "auto_total_score" in low_headers:
                score_col_idx = low_headers.index("auto_total_score") + 1
                score_col_letter = get_column_letter(score_col_idx)

                ws_low.conditional_formatting.add(
                    f"{score_col_letter}2:{score_col_letter}{ws_low.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=1,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=3,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=5,
                        end_color="63BE7B",
                    )
                )

        # =========================
        # 新增：图表分析 Sheet 图表
        # =========================

        ws_chart = wb["图表分析"]

        # 图表1：合格/不合格柱状图
        result_bar = BarChart()
        result_bar.title = "合格 / 不合格数量对比"
        result_bar.y_axis.title = "数量"
        result_bar.x_axis.title = "评估结果"

        result_data = Reference(ws_chart, min_col=2, min_row=4, max_row=6)
        result_cats = Reference(ws_chart, min_col=1, min_row=5, max_row=6)

        result_bar.add_data(result_data, titles_from_data=True)
        result_bar.set_categories(result_cats)
        result_bar.height = 7
        result_bar.width = 12

        ws_chart.add_chart(result_bar, "G3")

        # 图表2：合格/不合格饼图
        result_pie = PieChart()
        result_pie.title = "合格率分布"

        pie_data = Reference(ws_chart, min_col=2, min_row=4, max_row=6)
        pie_cats = Reference(ws_chart, min_col=1, min_row=5, max_row=6)

        result_pie.add_data(pie_data, titles_from_data=True)
        result_pie.set_categories(pie_cats)
        result_pie.height = 7
        result_pie.width = 10

        ws_chart.add_chart(result_pie, "M3")

        # 图表3：错误类型分布柱状图
        error_start_row = 5
        error_end_row = 4 + len(error_stats_df)

        if error_end_row >= error_start_row:
            error_bar = BarChart()
            error_bar.title = "错误类型分布"
            error_bar.y_axis.title = "数量"
            error_bar.x_axis.title = "错误类型"

            error_data = Reference(ws_chart, min_col=5, min_row=4, max_row=error_end_row)
            error_cats = Reference(ws_chart, min_col=4, min_row=5, max_row=error_end_row)

            error_bar.add_data(error_data, titles_from_data=True)
            error_bar.set_categories(error_cats)
            error_bar.height = 8
            error_bar.width = 14

            ws_chart.add_chart(error_bar, "G18")

        # 图表4：评分维度平均分柱状图
        score_start_row = 20
        score_end_row = 19 + len(score_stats_df)

        if score_end_row >= score_start_row:
            score_bar = BarChart()
            score_bar.title = "五项评分维度平均分"
            score_bar.y_axis.title = "平均分"
            score_bar.x_axis.title = "评分维度"

            score_data = Reference(ws_chart, min_col=2, min_row=19, max_row=score_end_row)
            score_cats = Reference(ws_chart, min_col=1, min_row=20, max_row=score_end_row)

            score_bar.add_data(score_data, titles_from_data=True)
            score_bar.set_categories(score_cats)
            score_bar.height = 8
            score_bar.width = 14

            ws_chart.add_chart(score_bar, "G34")

        # 图表5：低分样本错误类型柱状图
        low_error_start_row = 20
        low_error_end_row = 19 + len(low_error_stats_df)

        if low_error_end_row >= low_error_start_row:
            low_error_bar = BarChart()
            low_error_bar.title = "低分样本错误类型分布"
            low_error_bar.y_axis.title = "数量"
            low_error_bar.x_axis.title = "错误类型"

            low_error_data = Reference(ws_chart, min_col=5, min_row=19, max_row=low_error_end_row)
            low_error_cats = Reference(ws_chart, min_col=4, min_row=20, max_row=low_error_end_row)

            low_error_bar.add_data(low_error_data, titles_from_data=True)
            low_error_bar.set_categories(low_error_cats)
            low_error_bar.height = 8
            low_error_bar.width = 14

            ws_chart.add_chart(low_error_bar, "M18")

        # 图表分析页样式
        for col in range(1, 18):
            ws_chart.column_dimensions[get_column_letter(col)].width = 16

        for row in range(1, 50):
            ws_chart.row_dimensions[row].height = 22

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        for cell in ["A4", "B4", "D4", "E4", "A19", "B19", "D19", "E19"]:
            ws_chart[cell].fill = header_fill
            ws_chart[cell].font = Font(bold=True)
            ws_chart[cell].alignment = Alignment(horizontal="center", vertical="center")



        # =========================
        # 13. 调整 Sheet 顺序
        # =========================

        #sheet_order = ["汇总", "错误类型统计", "评分维度统计", "低分样本", "明细数据"]
        sheet_order = ["汇总", "图表分析", "错误类型统计", "评分维度统计", "低分样本", "明细数据"]
        wb._sheets = [wb[sheet_name] for sheet_name in sheet_order if sheet_name in wb.sheetnames]
