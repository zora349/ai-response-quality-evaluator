import re
from pathlib import Path

import pandas as pd

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_PATH = BASE_DIR / "data" / "auto_eval_samples.csv"
OUTPUT_PATH = BASE_DIR / "output" / "auto_eval_result.xlsx"
REPORT_PATH = BASE_DIR / "output" / "auto_eval_report.md"


def chinese_len(text: str) -> int:
    """中文场景下先简单按字符数统计。"""
    if pd.isna(text):
        return 0
    return len(str(text).strip())


def split_items(text: str):
    """
    将模型回答拆成条目。
    支持中文分号、英文分号、换行、句号。
    """
    if pd.isna(text):
        return []

    text = str(text).strip()
    separators = ["；", ";", "\n", "。"]

    items = [text]
    for sep in separators:
        new_items = []
        for item in items:
            new_items.extend(item.split(sep))
        items = new_items

    return [item.strip() for item in items if item.strip()]


def split_keywords(text: str):
    """
    将“卧室和小夜灯”“感应灯、磁吸”拆成关键词列表。
    """
    if not text:
        return []

    text = text.strip(" ：:“”\"'。；;，,")
    parts = re.split(r"[、,，和与及/]+", text)

    return [p.strip(" ：:“”\"'。；;，,") for p in parts if p.strip()]


def extract_rules(user_input: str):
    """
    从用户指令中动态提取规则。
    这是整个项目最核心的部分。
    """
    user_input = str(user_input)

    rules = {
        "item_count": None,
        "max_len_per_item": None,
        "max_len_total": None,
        "required_keywords": [],
        "forbidden_words": [],
        "no_explanation": False,
        "need_table": False,
        "need_simple": False,
        "no_exaggeration": False,
    }

    # 1. 提取条数：写3条 / 列出3点 / 生成5个 / 给我4条
    count_match = re.search(r"(?:写|列出|生成|给我|总结)?\s*(\d+)\s*(条|点|个)", user_input)
    if count_match:
        rules["item_count"] = int(count_match.group(1))

    # 2. 提取每条字数限制：每条不超过15字 / 每条10字以内 / 每条控制在12字内
    per_item_len_match = re.search(
        r"每条(?:不超过|控制在|限制在)?\s*(\d+)\s*字(?:以内|内)?",
        user_input
    )
    if per_item_len_match:
        rules["max_len_per_item"] = int(per_item_len_match.group(1))

    # 3. 提取整体字数限制：不超过30字 / 20字以内
    # 注意：如果已经存在“每条”限制，就不重复当成整体限制
    if "每条" not in user_input:
        total_len_match = re.search(
            r"(?:不超过|控制在|限制在)?\s*(\d+)\s*字(?:以内|内)",
            user_input
        )
        if total_len_match:
            rules["max_len_total"] = int(total_len_match.group(1))

    # 4. 提取必须包含关键词：必须包含卧室 / 必须包含感应灯和磁吸
    required_match = re.search(
        r"(?:必须包含|需要包含|标题包含|包含关键词)\s*([^，。；;,.]+)",
        user_input
    )
    if required_match:
        rules["required_keywords"] = split_keywords(required_match.group(1))

    # 5. 提取禁用词：不要出现新手 / 不能出现刺眼 / 禁止出现夸张词
    forbidden_matches = re.findall(
        r"(?:不要出现|不能出现|禁止出现|不要包含|不能包含)\s*([^，。；;,.]+)",
        user_input
    )
    for match in forbidden_matches:
        rules["forbidden_words"].extend(split_keywords(match))

    # 6. 不要解释
    if any(word in user_input for word in ["不要解释", "无需解释", "不用解释", "只给答案"]):
        rules["no_explanation"] = True

    # 7. 表格要求
    if any(word in user_input for word in ["用表格", "表格形式", "表格列出"]):
        rules["need_table"] = True

    # 8. 简单易懂要求
    if any(word in user_input for word in ["简单解释", "简单易懂", "通俗解释", "小白能懂"]):
        rules["need_simple"] = True

    # 9. 不要夸张语气
    if any(word in user_input for word in ["不要夸张", "不要使用夸张语气", "避免夸张"]):
        rules["no_exaggeration"] = True

    return rules


def check_item_count(model_output: str, rules: dict):
    expected = rules["item_count"]

    if expected is None:
        return True, "无条数要求"

    items = split_items(model_output)
    actual = len(items)

    if actual == expected:
        return True, f"条数符合要求：{actual}条"

    return False, f"条数不符合要求：要求{expected}条，实际{actual}条"


def check_length(model_output: str, rules: dict):
    items = split_items(model_output)

    max_len_per_item = rules["max_len_per_item"]
    max_len_total = rules["max_len_total"]

    if max_len_per_item is not None:
        over_items = [item for item in items if chinese_len(item) > max_len_per_item]

        if over_items:
            return False, f"存在超过{max_len_per_item}字的条目：{over_items}"

        return True, f"每条均不超过{max_len_per_item}字"

    if max_len_total is not None:
        actual_len = chinese_len(model_output)

        if actual_len > max_len_total:
            return False, f"整体字数超过{max_len_total}字，实际{actual_len}字"

        return True, f"整体字数不超过{max_len_total}字"

    return True, "无字数限制"


def check_required_keywords(model_output: str, rules: dict):
    required_keywords = rules["required_keywords"]

    if not required_keywords:
        return True, "无必须包含关键词"

    missing = [kw for kw in required_keywords if kw not in model_output]

    if missing:
        return False, f"缺少必须包含关键词：{missing}"

    return True, f"已包含必须关键词：{required_keywords}"


def check_forbidden_words(model_output: str, rules: dict):
    forbidden_words = rules["forbidden_words"]

    if not forbidden_words:
        return True, "无禁用词限制"

    appeared = [word for word in forbidden_words if word in model_output]

    if appeared:
        return False, f"出现禁用词：{appeared}"

    return True, f"未出现禁用词：{forbidden_words}"


def check_no_explanation(model_output: str, rules: dict):
    if not rules["no_explanation"]:
        return True, "无禁止解释要求"

    explanation_markers = [
        "这些方法",
        "以上",
        "可以帮助",
        "原因是",
        "因为",
        "建议你",
        "这样可以",
        "解释一下",
        "下面是",
    ]

    hit_words = [word for word in explanation_markers if word in model_output]

    if hit_words:
        return False, f"疑似存在多余解释：{hit_words}"

    return True, "未发现明显多余解释"


def check_table(model_output: str, rules: dict):
    if not rules["need_table"]:
        return True, "无表格要求"

    if "|" in model_output or "\t" in model_output:
        return True, "检测到表格结构"

    return False, "用户要求表格，但回答未体现表格结构"


def check_naturalness(user_input: str, model_output: str, rules: dict):
    """
    自然度很难纯规则判断，只做弱规则。
    更复杂的自然度适合人工或大模型判断。
    """
    if rules["no_exaggeration"]:
        exaggerated_words = [
            "超强",
            "秒杀",
            "必约",
            "爆款",
            "无敌",
            "绝对",
            "神器",
            "顶级",
        ]
        hit_words = [word for word in exaggerated_words if word in model_output]

        if hit_words:
            return False, f"存在夸张表达：{hit_words}"

    if rules["need_simple"]:
        complex_words = [
            "向量数据库",
            "embedding",
            "召回",
            "重排序",
            "系统化方法",
            "任务拆解",
            "上下文构建",
        ]
        hit_words = [word for word in complex_words if word in model_output]

        if hit_words:
            return False, f"表达偏复杂，不够简单易懂：{hit_words}"

    return True, "未发现明显自然度问题"


def chinese_num_to_int(text: str):
    """
    支持简单中文数字转阿拉伯数字。
    例如：一 -> 1，三 -> 3，十 -> 10
    """
    num_map = {
        "一": 1,
        "二": 2,
        "两": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }

    if text.isdigit():
        return int(text)

    return num_map.get(text)


def extract_count_from_text(text: str, prefix_keywords=None):
    """
    从文本中提取数量要求。
    支持：
    写3条
    写5个
    写2句
    列出三点
    生成4条
    模型只写2条
    模型写了两句
    """
    if prefix_keywords is None:
        prefix_keywords = ["写", "列出", "生成", "给出", "输出", "只写", "写了"]

    prefix_pattern = "|".join(prefix_keywords)

    pattern = rf"(?:{prefix_pattern})\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)"
    match = re.search(pattern, text)

    if not match:
        return None, None

    count = chinese_num_to_int(match.group(1))
    unit = match.group(2)

    return count, unit


def check_accuracy(user_input: str, model_output: str):
    """
    准确性检查：
    主要处理“判断下面回答是否合格”这类明确判断题。

    逻辑：
    1. 从用户问题中提取要求数量，例如：用户要求写3条
    2. 从用户问题中提取模型实际数量，例如：模型只写2条
    3. 如果实际数量少于要求数量，正确判断应为“不合格”
    4. 检查模型回答是否给出了正确判断
    """
    user_input = str(user_input)
    model_output = str(model_output)

    if "判断下面回答是否合格" not in user_input:
        return True, "非判断题，准确性需结合具体语义判断"

    # 提取“用户要求写3条”
    required_match = re.search(
        r"用户要求(?:写|列出|生成|给出|输出)\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)",
        user_input
    )

    # 提取“模型只写2条 / 模型写了2条”
    actual_match = re.search(
        r"模型(?:只写|写了|只列出|列出|生成了)?\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)",
        user_input
    )

    if required_match and actual_match:
        required_count = chinese_num_to_int(required_match.group(1))
        required_unit = required_match.group(2)

        actual_count = chinese_num_to_int(actual_match.group(1))
        actual_unit = actual_match.group(2)

        # 单位一致，且数量不足
        if required_unit == actual_unit and actual_count < required_count:
            if "不合格" in model_output:
                return True, f"判断正确：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，应为不合格"

            return False, f"判断错误：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，应判为不合格"

        # 单位一致，且数量满足
        if required_unit == actual_unit and actual_count >= required_count:
            if "合格" in model_output and "不合格" not in model_output:
                return True, f"判断正确：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，数量满足"

            return False, f"判断错误：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，数量满足，应判为合格"

    return True, "未提取到明确数量判断规则，准确性需人工复核"


def score_from_checks(checks: dict):
    """
    根据检查结果生成五项评分。
    """
    instruction_related = [
        "required_keywords",
        "forbidden_words",
        "no_explanation",
    ]

    completeness_related = [
        "item_count",
    ]

    format_related = [
        "length",
        "table",
        "item_count",
    ]

    naturalness_related = [
        "naturalness",
    ]

    accuracy_related = [
        "accuracy",
    ]

    def calc_score(keys):
        failed_count = sum(1 for k in keys if checks[k][0] is False)

        if failed_count == 0:
            return 5
        if failed_count == 1:
            return 2
        return 1

    return {
        "auto_instruction_score": calc_score(instruction_related),
        "auto_completeness_score": calc_score(completeness_related),
        "auto_format_score": calc_score(format_related),
        "auto_naturalness_score": calc_score(naturalness_related),
        "auto_accuracy_score": calc_score(accuracy_related),
    }


def classify_error(checks: dict):
    failed_reasons = [
        reason for passed, reason in checks.values()
        if passed is False
    ]

    reason_text = "；".join(failed_reasons)

    if not failed_reasons:
        return "无明显错误"

    if "判断错误" in reason_text:
        return "判断错误"

    if "缺少必须包含关键词" in reason_text:
        return "指令违背"

    if "出现禁用词" in reason_text:
        return "指令违背"

    if "多余解释" in reason_text:
        return "指令违背"

    if "条数不符合" in reason_text:
        return "回答不完整"

    if "超过" in reason_text or "表格" in reason_text:
        return "格式/长度错误"

    if "夸张表达" in reason_text or "不够简单易懂" in reason_text:
        return "表达不符合要求"

    return "其他错误"


def evaluate_row(row):
    '''
    ---------------------------------------------
    每次拿 CSV 表格里的一行数据，自动完成一次完整评估。

    1. 读取这一行的 user_input 和 model_output
    2. 从 user_input 里解析规则
    3. 用规则检查 model_output
    4. 给这条回答打五项评分
    5. 判断合格/不合格
    6. 归因错误类型
    7. 输出判断理由
    -------------------------------------------
    '''
    user_input = str(row["user_input"])
    model_output = str(row["model_output"])

    rules = extract_rules(user_input)

    checks = {
        "item_count": check_item_count(model_output, rules),
        "length": check_length(model_output, rules),
        "required_keywords": check_required_keywords(model_output, rules),
        "forbidden_words": check_forbidden_words(model_output, rules),
        "no_explanation": check_no_explanation(model_output, rules),
        "table": check_table(model_output, rules),
        "naturalness": check_naturalness(user_input, model_output, rules),
        "accuracy": check_accuracy(user_input, model_output),
    }

    scores = score_from_checks(checks)

    total_score = round(sum(scores.values()) / len(scores), 2)

    failed_reasons = [
        reason for passed, reason in checks.values()
        if passed is False
    ]

    if failed_reasons:
        reason = "；".join(failed_reasons)
    else:
        reason = "；".join(reason for passed, reason in checks.values())

    error_type = classify_error(checks)

    if (
        total_score >= 4
        and scores["auto_instruction_score"] >= 4
        and scores["auto_format_score"] >= 4
    ):
        auto_result = "合格"
    else:
        auto_result = "不合格"

    return pd.Series({
        "parsed_rules": str(rules),
        **scores,
        "auto_total_score": total_score,
        "auto_result": auto_result,
        "auto_error_type": error_type,
        "auto_reason": reason,
    })

def generate_excel_analysis(result_df: pd.DataFrame, output_path: Path):
    """
    生成多Sheet Excel分析报告：
    1. 汇总
    2. 错误类型统计
    3. 评分维度统计
    4. 低分样本
    5. 明细数据

    同时生成柱状图/饼图，并做基础样式美化。
    """

    # =========================
    # 1. 统计基础数据
    # =========================

    total = len(result_df)
    passed = len(result_df[result_df["auto_result"] == "合格"])
    failed = len(result_df[result_df["auto_result"] == "不合格"])
    pass_rate = passed / total if total else 0

    score_columns = [
        "auto_instruction_score",
        "auto_completeness_score",
        "auto_format_score",
        "auto_naturalness_score",
        "auto_accuracy_score",
        "auto_total_score",
    ]

    score_name_map = {
        "auto_instruction_score": "指令遵循",
        "auto_completeness_score": "回答完整性",
        "auto_format_score": "格式规范",
        "auto_naturalness_score": "表达自然度",
        "auto_accuracy_score": "准确性",
        "auto_total_score": "综合得分",
    }

    # =========================
    # 2. 汇总 Sheet 数据
    # =========================

    summary_df = pd.DataFrame({
        "指标": [
            "总样本数",
            "合格数量",
            "不合格数量",
            "合格率",
            "综合平均分",
            "指令遵循平均分",
            "回答完整性平均分",
            "格式规范平均分",
            "表达自然度平均分",
            "准确性平均分",
        ],
        "数值": [
            total,
            passed,
            failed,
            pass_rate,
            round(result_df["auto_total_score"].mean(), 2),
            round(result_df["auto_instruction_score"].mean(), 2),
            round(result_df["auto_completeness_score"].mean(), 2),
            round(result_df["auto_format_score"].mean(), 2),
            round(result_df["auto_naturalness_score"].mean(), 2),
            round(result_df["auto_accuracy_score"].mean(), 2),
        ]
    })

    result_distribution_df = pd.DataFrame({
        "结果": ["合格", "不合格"],
        "数量": [passed, failed],
    })

    # =========================
    # 3. 错误类型统计
    # =========================

    error_stats_df = (
        result_df["auto_error_type"]
        .value_counts()
        .reset_index()
    )
    error_stats_df.columns = ["错误类型", "数量"]
    error_stats_df["占比"] = error_stats_df["数量"] / total

    # =========================
    # 4. 评分维度统计
    # =========================

    score_stats_list = []

    for col in score_columns:
        score_stats_list.append({
            "评分维度": score_name_map[col],
            "平均分": round(result_df[col].mean(), 2),
            "最低分": result_df[col].min(),
            "最高分": result_df[col].max(),
        })

    score_stats_df = pd.DataFrame(score_stats_list)

    # =========================
    # 5. 低分样本
    # =========================

    low_score_df = result_df[
        (result_df["auto_result"] == "不合格") |
        (result_df["auto_total_score"] < 4)
    ].copy()

    low_score_columns = [
        "id",
        "user_input",
        "model_output",
        "parsed_rules",
        "auto_instruction_score",
        "auto_completeness_score",
        "auto_format_score",
        "auto_naturalness_score",
        "auto_accuracy_score",
        "auto_total_score",
        "auto_result",
        "auto_error_type",
        "auto_reason",
    ]

    low_score_columns = [col for col in low_score_columns if col in low_score_df.columns]
    low_score_df = low_score_df[low_score_columns].sort_values(
        by="auto_total_score",
        ascending=True
    )

    # =========================
    # 6. 写入 Excel 多个 Sheet
    # =========================

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总", index=False, startrow=0)
        result_distribution_df.to_excel(writer, sheet_name="汇总", index=False, startrow=13)

        error_stats_df.to_excel(writer, sheet_name="错误类型统计", index=False)
        score_stats_df.to_excel(writer, sheet_name="评分维度统计", index=False)
        low_score_df.to_excel(writer, sheet_name="低分样本", index=False)
        result_df.to_excel(writer, sheet_name="明细数据", index=False)

        wb = writer.book


        # =========================
        # 新增：图表分析 Sheet
        # =========================

        ws_chart = wb.create_sheet("图表分析")

        ws_chart["A1"] = "AI回复质量自动评估图表分析"
        ws_chart["A1"].font = Font(bold=True, size=16)

        # 1. 合格/不合格分布数据
        ws_chart["A3"] = "评估结果分布"
        ws_chart["A3"].font = Font(bold=True, size=12)

        ws_chart["A4"] = "结果"
        ws_chart["B4"] = "数量"
        ws_chart["A5"] = "合格"
        ws_chart["B5"] = passed
        ws_chart["A6"] = "不合格"
        ws_chart["B6"] = failed

        # 2. 错误类型统计数据
        ws_chart["D3"] = "错误类型分布"
        ws_chart["D3"].font = Font(bold=True, size=12)

        ws_chart["D4"] = "错误类型"
        ws_chart["E4"] = "数量"

        for idx, row in error_stats_df.iterrows():
            excel_row = idx + 5
            ws_chart[f"D{excel_row}"] = row["错误类型"]
            ws_chart[f"E{excel_row}"] = row["数量"]

        # 3. 评分维度统计数据
        ws_chart["A18"] = "评分维度平均分"
        ws_chart["A18"].font = Font(bold=True, size=12)

        ws_chart["A19"] = "评分维度"
        ws_chart["B19"] = "平均分"

        for idx, row in score_stats_df.iterrows():
            excel_row = idx + 20
            ws_chart[f"A{excel_row}"] = row["评分维度"]
            ws_chart[f"B{excel_row}"] = row["平均分"]

        # 4. 低分样本错误类型统计
        ws_chart["D18"] = "低分样本错误类型"
        ws_chart["D18"].font = Font(bold=True, size=12)

        low_error_stats_df = (
            low_score_df["auto_error_type"]
            .value_counts()
            .reset_index()
        )

        low_error_stats_df.columns = ["错误类型", "数量"]

        ws_chart["D19"] = "错误类型"
        ws_chart["E19"] = "数量"

        for idx, row in low_error_stats_df.iterrows():
            excel_row = idx + 20
            ws_chart[f"D{excel_row}"] = row["错误类型"]
            ws_chart[f"E{excel_row}"] = row["数量"]

        # =========================
        # 7. 通用样式函数
        # =========================

        def style_worksheet(ws):
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            header_font = Font(bold=True, color="000000")
            thin_side = Side(style="thin", color="D9D9D9")
            border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )

            ws.freeze_panes = "A2"

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=True
                    )
                    cell.border = border

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # 自动列宽，但限制最大宽度，避免太夸张
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_length = 0

                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max(max_length + 2, 10), 45)
                ws.column_dimensions[col_letter].width = adjusted_width

            # 行高
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 22

        for sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name])

        # =========================
        # 8. 汇总 Sheet 样式 + 图表
        # =========================

        ws_summary = wb["汇总"]

        ws_summary["D1"] = "合格/不合格分布"
        ws_summary["D1"].font = Font(bold=True, size=12)

        bar_chart = BarChart()
        bar_chart.title = "合格 / 不合格数量"
        bar_chart.y_axis.title = "数量"
        bar_chart.x_axis.title = "结果"

        data = Reference(ws_summary, min_col=2, min_row=14, max_row=16)
        cats = Reference(ws_summary, min_col=1, min_row=15, max_row=16)

        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.height = 7
        bar_chart.width = 12

        ws_summary.add_chart(bar_chart, "D2")

        pie_chart = PieChart()
        pie_chart.title = "合格率分布"

        pie_data = Reference(ws_summary, min_col=2, min_row=14, max_row=16)
        pie_cats = Reference(ws_summary, min_col=1, min_row=15, max_row=16)

        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_cats)
        pie_chart.height = 7
        pie_chart.width = 10

        ws_summary.add_chart(pie_chart, "D18")

        # 合格率格式
        ws_summary["B5"].number_format = "0.00%"

        # =========================
        # 9. 错误类型统计 Sheet 图表
        # =========================

        ws_error = wb["错误类型统计"]

        if ws_error.max_row >= 2:
            error_chart = BarChart()
            error_chart.title = "错误类型分布"
            error_chart.y_axis.title = "数量"
            error_chart.x_axis.title = "错误类型"

            error_data = Reference(ws_error, min_col=2, min_row=1, max_row=ws_error.max_row)
            error_cats = Reference(ws_error, min_col=1, min_row=2, max_row=ws_error.max_row)

            error_chart.add_data(error_data, titles_from_data=True)
            error_chart.set_categories(error_cats)
            error_chart.height = 8
            error_chart.width = 14

            ws_error.add_chart(error_chart, "E2")

            # 占比列格式
            for row in range(2, ws_error.max_row + 1):
                ws_error[f"C{row}"].number_format = "0.00%"

        # =========================
        # 10. 评分维度统计 Sheet 图表
        # =========================

        ws_score = wb["评分维度统计"]

        if ws_score.max_row >= 2:
            score_chart = BarChart()
            score_chart.title = "五项评分维度平均分"
            score_chart.y_axis.title = "平均分"
            score_chart.x_axis.title = "评分维度"

            score_data = Reference(ws_score, min_col=2, min_row=1, max_row=ws_score.max_row)
            score_cats = Reference(ws_score, min_col=1, min_row=2, max_row=ws_score.max_row)

            score_chart.add_data(score_data, titles_from_data=True)
            score_chart.set_categories(score_cats)
            score_chart.height = 8
            score_chart.width = 14

            ws_score.add_chart(score_chart, "F2")

        # =========================
        # 11. 明细数据 Sheet 条件格式
        # =========================

        ws_detail = wb["明细数据"]

        # 找到评分列位置
        detail_headers = [cell.value for cell in ws_detail[1]]

        for score_col in score_columns:
            if score_col in detail_headers:
                col_idx = detail_headers.index(score_col) + 1
                col_letter = get_column_letter(col_idx)

                ws_detail.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws_detail.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=1,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=3,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=5,
                        end_color="63BE7B",
                    )
                )

        # =========================
        # 12. 低分样本 Sheet 突出低分
        # =========================

        ws_low = wb["低分样本"]

        if ws_low.max_row >= 2:
            low_headers = [cell.value for cell in ws_low[1]]

            if "auto_total_score" in low_headers:
                score_col_idx = low_headers.index("auto_total_score") + 1
                score_col_letter = get_column_letter(score_col_idx)

                ws_low.conditional_formatting.add(
                    f"{score_col_letter}2:{score_col_letter}{ws_low.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=1,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=3,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=5,
                        end_color="63BE7B",
                    )
                )

        # =========================
        # 新增：图表分析 Sheet 图表
        # =========================

        ws_chart = wb["图表分析"]

        # 图表1：合格/不合格柱状图
        result_bar = BarChart()
        result_bar.title = "合格 / 不合格数量对比"
        result_bar.y_axis.title = "数量"
        result_bar.x_axis.title = "评估结果"

        result_data = Reference(ws_chart, min_col=2, min_row=4, max_row=6)
        result_cats = Reference(ws_chart, min_col=1, min_row=5, max_row=6)

        result_bar.add_data(result_data, titles_from_data=True)
        result_bar.set_categories(result_cats)
        result_bar.height = 7
        result_bar.width = 12

        ws_chart.add_chart(result_bar, "G3")

        # 图表2：合格/不合格饼图
        result_pie = PieChart()
        result_pie.title = "合格率分布"

        pie_data = Reference(ws_chart, min_col=2, min_row=4, max_row=6)
        pie_cats = Reference(ws_chart, min_col=1, min_row=5, max_row=6)

        result_pie.add_data(pie_data, titles_from_data=True)
        result_pie.set_categories(pie_cats)
        result_pie.height = 7
        result_pie.width = 10

        ws_chart.add_chart(result_pie, "M3")

        # 图表3：错误类型分布柱状图
        error_start_row = 5
        error_end_row = 4 + len(error_stats_df)

        if error_end_row >= error_start_row:
            error_bar = BarChart()
            error_bar.title = "错误类型分布"
            error_bar.y_axis.title = "数量"
            error_bar.x_axis.title = "错误类型"

            error_data = Reference(ws_chart, min_col=5, min_row=4, max_row=error_end_row)
            error_cats = Reference(ws_chart, min_col=4, min_row=5, max_row=error_end_row)

            error_bar.add_data(error_data, titles_from_data=True)
            error_bar.set_categories(error_cats)
            error_bar.height = 8
            error_bar.width = 14

            ws_chart.add_chart(error_bar, "G18")

        # 图表4：评分维度平均分柱状图
        score_start_row = 20
        score_end_row = 19 + len(score_stats_df)

        if score_end_row >= score_start_row:
            score_bar = BarChart()
            score_bar.title = "五项评分维度平均分"
            score_bar.y_axis.title = "平均分"
            score_bar.x_axis.title = "评分维度"

            score_data = Reference(ws_chart, min_col=2, min_row=19, max_row=score_end_row)
            score_cats = Reference(ws_chart, min_col=1, min_row=20, max_row=score_end_row)

            score_bar.add_data(score_data, titles_from_data=True)
            score_bar.set_categories(score_cats)
            score_bar.height = 8
            score_bar.width = 14

            ws_chart.add_chart(score_bar, "G34")

        # 图表5：低分样本错误类型柱状图
        low_error_start_row = 20
        low_error_end_row = 19 + len(low_error_stats_df)

        if low_error_end_row >= low_error_start_row:
            low_error_bar = BarChart()
            low_error_bar.title = "低分样本错误类型分布"
            low_error_bar.y_axis.title = "数量"
            low_error_bar.x_axis.title = "错误类型"

            low_error_data = Reference(ws_chart, min_col=5, min_row=19, max_row=low_error_end_row)
            low_error_cats = Reference(ws_chart, min_col=4, min_row=20, max_row=low_error_end_row)

            low_error_bar.add_data(low_error_data, titles_from_data=True)
            low_error_bar.set_categories(low_error_cats)
            low_error_bar.height = 8
            low_error_bar.width = 14

            ws_chart.add_chart(low_error_bar, "M18")

        # 图表分析页样式
        for col in range(1, 18):
            ws_chart.column_dimensions[get_column_letter(col)].width = 16

        for row in range(1, 50):
            ws_chart.row_dimensions[row].height = 22

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        for cell in ["A4", "B4", "D4", "E4", "A19", "B19", "D19", "E19"]:
            ws_chart[cell].fill = header_fill
            ws_chart[cell].font = Font(bold=True)
            ws_chart[cell].alignment = Alignment(horizontal="center", vertical="center")



        # =========================
        # 13. 调整 Sheet 顺序
        # =========================

        #sheet_order = ["汇总", "错误类型统计", "评分维度统计", "低分样本", "明细数据"]
        sheet_order = ["汇总", "图表分析", "错误类型统计", "评分维度统计", "低分样本", "明细数据"]
        wb._sheets = [wb[sheet_name] for sheet_name in sheet_order if sheet_name in wb.sheetnames]


def generate_report(df):
    total = len(df)
    passed = len(df[df["auto_result"] == "合格"])
    failed = len(df[df["auto_result"] == "不合格"])
    pass_rate = passed / total * 100 if total else 0

    report = []

    report.append("# AI回复质量自动评估报告\n")

    report.append("## 一、项目概况\n")
    report.append(
        "本项目基于 Python 实现 AI 回复质量自动评估。"
        "系统从用户指令中动态解析条数、字数、必须包含关键词、禁用词、表格要求、禁止解释等规则，"
        "再对模型回复进行多维度评分和错误归因。\n"
    )

    report.append("## 二、整体结果\n")
    report.append(f"- 总样本数：{total}\n")
    report.append(f"- 合格数量：{passed}\n")
    report.append(f"- 不合格数量：{failed}\n")
    report.append(f"- 合格率：{pass_rate:.2f}%\n")
    report.append(f"- 综合平均分：{df['auto_total_score'].mean():.2f}\n")

    report.append("## 三、五项指标平均分\n")
    report.append(f"- 指令遵循平均分：{df['auto_instruction_score'].mean():.2f}\n")
    report.append(f"- 完整性平均分：{df['auto_completeness_score'].mean():.2f}\n")
    report.append(f"- 格式规范平均分：{df['auto_format_score'].mean():.2f}\n")
    report.append(f"- 自然度平均分：{df['auto_naturalness_score'].mean():.2f}\n")
    report.append(f"- 准确性平均分：{df['auto_accuracy_score'].mean():.2f}\n")

    report.append("## 四、错误类型统计\n")
    for error_type, count in df["auto_error_type"].value_counts().items():
        report.append(f"- {error_type}：{count}条\n")

    report.append("## 五、典型错误样本\n")
    failed_df = df[df["auto_result"] == "不合格"]

    for _, row in failed_df.iterrows():
        report.append(f"### 样本 {row['id']}\n")
        report.append(f"- 用户要求：{row['user_input']}\n")
        report.append(f"- 模型回答：{row['model_output']}\n")
        report.append(f"- 解析规则：{row['parsed_rules']}\n")
        report.append(f"- 综合得分：{row['auto_total_score']}\n")
        report.append(f"- 错误类型：{row['auto_error_type']}\n")
        report.append(f"- 判断理由：{row['auto_reason']}\n")

    report.append("## 六、优化建议\n")
    report.append("1. 对指令违背类问题，应加强模型对必须包含、禁用词、禁止解释等限制条件的遵循能力。\n")
    report.append("2. 对格式/长度错误，应增加条数、字数、表格格式等约束类训练样本。\n")
    report.append("3. 对回答不完整问题，应优化模型对多条件任务的拆解能力。\n")
    report.append("4. 对自然度、准确性等语义类指标，建议结合人工复核或大模型辅助评估。\n")

    REPORT_PATH.write_text("\n".join(report), encoding="utf-8")


def main():
    df = pd.read_csv(DATA_PATH)

    eval_result = df.apply(evaluate_row, axis=1)
    result_df = pd.concat([df, eval_result], axis=1)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    #result_df.to_excel(OUTPUT_PATH, index=False)
    generate_excel_analysis(result_df, OUTPUT_PATH)
    generate_report(result_df)

    print("评估完成")
    print(f"总样本数: {len(result_df)}")
    print(f"合格数量: {len(result_df[result_df['auto_result'] == '合格'])}")
    print(f"不合格数量: {len(result_df[result_df['auto_result'] == '不合格'])}")
    print(f"综合平均分: {result_df['auto_total_score'].mean():.2f}")
    print(f"Excel结果已保存: {OUTPUT_PATH}")
    print(f"项目报告已保存: {REPORT_PATH}")


if __name__ == "__main__":
    main()