import re

import pandas as pd

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from config import REPORT_PATH

def generate_report(df):
    """

    汇总报告生成模块

    负责基于评估结果生成 AI 回复质量自动评估报告，
    包括整体合格率、平均分、错误类型统计和质量分析总结。

    """

    total = len(df)
    passed = len(df[df["auto_result"] == "合格"])
    failed = len(df[df["auto_result"] == "不合格"])
    pass_rate = passed / total * 100 if total else 0

    report = []

    report.append("# AI回复质量自动评估报告\n")

    report.append("## 一、项目概况\n")
    report.append(
        "本项目基于 Python 实现 AI 回复质量自动评估。"
        "系统从用户指令中动态解析条数、字数、必须包含关键词、禁用词、表格要求、禁止解释等规则，"
        "再对模型回复进行多维度评分和错误归因。\n"
    )

    report.append("## 二、整体结果\n")
    report.append(f"- 总样本数：{total}\n")
    report.append(f"- 合格数量：{passed}\n")
    report.append(f"- 不合格数量：{failed}\n")
    report.append(f"- 合格率：{pass_rate:.2f}%\n")
    report.append(f"- 综合平均分：{df['auto_total_score'].mean():.2f}\n")

    report.append("## 三、五项指标平均分\n")
    report.append(f"- 指令遵循平均分：{df['auto_instruction_score'].mean():.2f}\n")
    report.append(f"- 完整性平均分：{df['auto_completeness_score'].mean():.2f}\n")
    report.append(f"- 格式规范平均分：{df['auto_format_score'].mean():.2f}\n")
    report.append(f"- 自然度平均分：{df['auto_naturalness_score'].mean():.2f}\n")
    report.append(f"- 准确性平均分：{df['auto_accuracy_score'].mean():.2f}\n")

    report.append("## 四、错误类型统计\n")
    for error_type, count in df["auto_error_type"].value_counts().items():
        report.append(f"- {error_type}：{count}条\n")

    report.append("## 五、典型错误样本\n")
    failed_df = df[df["auto_result"] == "不合格"]

    for _, row in failed_df.iterrows():
        report.append(f"### 样本 {row['id']}\n")
        report.append(f"- 用户要求：{row['user_input']}\n")
        report.append(f"- 模型回答：{row['model_output']}\n")
        report.append(f"- 解析规则：{row['parsed_rules']}\n")
        report.append(f"- 综合得分：{row['auto_total_score']}\n")
        report.append(f"- 错误类型：{row['auto_error_type']}\n")
        report.append(f"- 判断理由：{row['auto_reason']}\n")

    report.append("## 六、优化建议\n")
    report.append("1. 对指令违背类问题，应加强模型对必须包含、禁用词、禁止解释等限制条件的遵循能力。\n")
    report.append("2. 对格式/长度错误，应增加条数、字数、表格格式等约束类训练样本。\n")
    report.append("3. 对回答不完整问题，应优化模型对多条件任务的拆解能力。\n")
    report.append("4. 对自然度、准确性等语义类指标，建议结合人工复核或大模型辅助评估。\n")

    REPORT_PATH.write_text("\n".join(report), encoding="utf-8")