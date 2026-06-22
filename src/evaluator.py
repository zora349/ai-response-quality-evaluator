import re

import pandas as pd

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

def chinese_len(text: str) -> int:
    """中文场景下先简单按字符数统计。"""
    if pd.isna(text):
        return 0
    return len(str(text).strip())


def split_items(text: str):
    """
    将模型回答拆成条目。
    支持中文分号、英文分号、换行、句号。
    """
    if pd.isna(text):
        return []

    text = str(text).strip()
    separators = ["；", ";", "\n", "。"]

    items = [text]
    for sep in separators:
        new_items = []
        for item in items:
            new_items.extend(item.split(sep))
        items = new_items

    return [item.strip() for item in items if item.strip()]


def split_keywords(text: str):
    """
    将“卧室和小夜灯”“感应灯、磁吸”拆成关键词列表。
    """
    if not text:
        return []

    text = text.strip(" ：:“”\"'。；;，,")
    parts = re.split(r"[、,，和与及/]+", text)

    return [p.strip(" ：:“”\"'。；;，,") for p in parts if p.strip()]


def extract_rules(user_input: str):
    """
    从用户指令中动态提取规则。
    这是整个项目最核心的部分。
    """
    user_input = str(user_input)

    rules = {
        "item_count": None,
        "max_len_per_item": None,
        "max_len_total": None,
        "required_keywords": [],
        "forbidden_words": [],
        "no_explanation": False,
        "need_table": False,
        "need_simple": False,
        "no_exaggeration": False,
    }

    # 1. 提取条数：写3条 / 列出3点 / 生成5个 / 给我4条
    count_match = re.search(r"(?:写|列出|生成|给我|总结)?\s*(\d+)\s*(条|点|个)", user_input)
    if count_match:
        rules["item_count"] = int(count_match.group(1))

    # 2. 提取每条字数限制：每条不超过15字 / 每条10字以内 / 每条控制在12字内
    per_item_len_match = re.search(
        r"每条(?:不超过|控制在|限制在)?\s*(\d+)\s*字(?:以内|内)?",
        user_input
    )
    if per_item_len_match:
        rules["max_len_per_item"] = int(per_item_len_match.group(1))

    # 3. 提取整体字数限制：不超过30字 / 20字以内
    # 注意：如果已经存在“每条”限制，就不重复当成整体限制
    if "每条" not in user_input:
        total_len_match = re.search(
            r"(?:不超过|控制在|限制在)?\s*(\d+)\s*字(?:以内|内)",
            user_input
        )
        if total_len_match:
            rules["max_len_total"] = int(total_len_match.group(1))

    # 4. 提取必须包含关键词：必须包含卧室 / 必须包含感应灯和磁吸
    required_match = re.search(
        r"(?:必须包含|需要包含|标题包含|包含关键词)\s*([^，。；;,.]+)",
        user_input
    )
    if required_match:
        rules["required_keywords"] = split_keywords(required_match.group(1))

    # 5. 提取禁用词：不要出现新手 / 不能出现刺眼 / 禁止出现夸张词
    forbidden_matches = re.findall(
        r"(?:不要出现|不能出现|禁止出现|不要包含|不能包含)\s*([^，。；;,.]+)",
        user_input
    )
    for match in forbidden_matches:
        rules["forbidden_words"].extend(split_keywords(match))

    # 6. 不要解释
    if any(word in user_input for word in ["不要解释", "无需解释", "不用解释", "只给答案"]):
        rules["no_explanation"] = True

    # 7. 表格要求
    if any(word in user_input for word in ["用表格", "表格形式", "表格列出"]):
        rules["need_table"] = True

    # 8. 简单易懂要求
    if any(word in user_input for word in ["简单解释", "简单易懂", "通俗解释", "小白能懂"]):
        rules["need_simple"] = True

    # 9. 不要夸张语气
    if any(word in user_input for word in ["不要夸张", "不要使用夸张语气", "避免夸张"]):
        rules["no_exaggeration"] = True

    return rules


def check_item_count(model_output: str, rules: dict):
    expected = rules["item_count"]

    if expected is None:
        return True, "无条数要求"

    items = split_items(model_output)
    actual = len(items)

    if actual == expected:
        return True, f"条数符合要求：{actual}条"

    return False, f"条数不符合要求：要求{expected}条，实际{actual}条"


def check_length(model_output: str, rules: dict):
    items = split_items(model_output)

    max_len_per_item = rules["max_len_per_item"]
    max_len_total = rules["max_len_total"]

    if max_len_per_item is not None:
        over_items = [item for item in items if chinese_len(item) > max_len_per_item]

        if over_items:
            return False, f"存在超过{max_len_per_item}字的条目：{over_items}"

        return True, f"每条均不超过{max_len_per_item}字"

    if max_len_total is not None:
        actual_len = chinese_len(model_output)

        if actual_len > max_len_total:
            return False, f"整体字数超过{max_len_total}字，实际{actual_len}字"

        return True, f"整体字数不超过{max_len_total}字"

    return True, "无字数限制"


def check_required_keywords(model_output: str, rules: dict):
    required_keywords = rules["required_keywords"]

    if not required_keywords:
        return True, "无必须包含关键词"

    missing = [kw for kw in required_keywords if kw not in model_output]

    if missing:
        return False, f"缺少必须包含关键词：{missing}"

    return True, f"已包含必须关键词：{required_keywords}"


def check_forbidden_words(model_output: str, rules: dict):
    forbidden_words = rules["forbidden_words"]

    if not forbidden_words:
        return True, "无禁用词限制"

    appeared = [word for word in forbidden_words if word in model_output]

    if appeared:
        return False, f"出现禁用词：{appeared}"

    return True, f"未出现禁用词：{forbidden_words}"


def check_no_explanation(model_output: str, rules: dict):
    if not rules["no_explanation"]:
        return True, "无禁止解释要求"

    explanation_markers = [
        "这些方法",
        "以上",
        "可以帮助",
        "原因是",
        "因为",
        "建议你",
        "这样可以",
        "解释一下",
        "下面是",
    ]

    hit_words = [word for word in explanation_markers if word in model_output]

    if hit_words:
        return False, f"疑似存在多余解释：{hit_words}"

    return True, "未发现明显多余解释"


def check_table(model_output: str, rules: dict):
    if not rules["need_table"]:
        return True, "无表格要求"

    if "|" in model_output or "\t" in model_output:
        return True, "检测到表格结构"

    return False, "用户要求表格，但回答未体现表格结构"


def check_naturalness(user_input: str, model_output: str, rules: dict):
    """
    自然度很难纯规则判断，只做弱规则。
    更复杂的自然度适合人工或大模型判断。
    """
    if rules["no_exaggeration"]:
        exaggerated_words = [
            "超强",
            "秒杀",
            "必约",
            "爆款",
            "无敌",
            "绝对",
            "神器",
            "顶级",
        ]
        hit_words = [word for word in exaggerated_words if word in model_output]

        if hit_words:
            return False, f"存在夸张表达：{hit_words}"

    if rules["need_simple"]:
        complex_words = [
            "向量数据库",
            "embedding",
            "召回",
            "重排序",
            "系统化方法",
            "任务拆解",
            "上下文构建",
        ]
        hit_words = [word for word in complex_words if word in model_output]

        if hit_words:
            return False, f"表达偏复杂，不够简单易懂：{hit_words}"

    return True, "未发现明显自然度问题"


def chinese_num_to_int(text: str):
    """
    支持简单中文数字转阿拉伯数字。
    例如：一 -> 1，三 -> 3，十 -> 10
    """
    num_map = {
        "一": 1,
        "二": 2,
        "两": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }

    if text.isdigit():
        return int(text)

    return num_map.get(text)


def extract_count_from_text(text: str, prefix_keywords=None):
    """
    从文本中提取数量要求。
    支持：
    写3条
    写5个
    写2句
    列出三点
    生成4条
    模型只写2条
    模型写了两句
    """
    if prefix_keywords is None:
        prefix_keywords = ["写", "列出", "生成", "给出", "输出", "只写", "写了"]

    prefix_pattern = "|".join(prefix_keywords)

    pattern = rf"(?:{prefix_pattern})\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)"
    match = re.search(pattern, text)

    if not match:
        return None, None

    count = chinese_num_to_int(match.group(1))
    unit = match.group(2)

    return count, unit


def check_accuracy(user_input: str, model_output: str):
    """
    准确性检查：
    主要处理“判断下面回答是否合格”这类明确判断题。

    逻辑：
    1. 从用户问题中提取要求数量，例如：用户要求写3条
    2. 从用户问题中提取模型实际数量，例如：模型只写2条
    3. 如果实际数量少于要求数量，正确判断应为“不合格”
    4. 检查模型回答是否给出了正确判断
    """
    user_input = str(user_input)
    model_output = str(model_output)

    if "判断下面回答是否合格" not in user_input:
        return True, "非判断题，准确性需结合具体语义判断"

    # 提取“用户要求写3条”
    required_match = re.search(
        r"用户要求(?:写|列出|生成|给出|输出)\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)",
        user_input
    )

    # 提取“模型只写2条 / 模型写了2条”
    actual_match = re.search(
        r"模型(?:只写|写了|只列出|列出|生成了)?\s*([0-9]+|[一二两三四五六七八九十])\s*(条|个|句|点|项)",
        user_input
    )

    if required_match and actual_match:
        required_count = chinese_num_to_int(required_match.group(1))
        required_unit = required_match.group(2)

        actual_count = chinese_num_to_int(actual_match.group(1))
        actual_unit = actual_match.group(2)

        # 单位一致，且数量不足
        if required_unit == actual_unit and actual_count < required_count:
            if "不合格" in model_output:
                return True, f"判断正确：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，应为不合格"

            return False, f"判断错误：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，应判为不合格"

        # 单位一致，且数量满足
        if required_unit == actual_unit and actual_count >= required_count:
            if "合格" in model_output and "不合格" not in model_output:
                return True, f"判断正确：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，数量满足"

            return False, f"判断错误：要求{required_count}{required_unit}，实际{actual_count}{actual_unit}，数量满足，应判为合格"

    return True, "未提取到明确数量判断规则，准确性需人工复核"

def score_from_checks(checks: dict):
    """
    根据检查结果生成五项评分。
    """
    instruction_related = [
        "required_keywords",
        "forbidden_words",
        "no_explanation",
    ]

    completeness_related = [
        "item_count",
    ]

    format_related = [
        "length",
        "table",
        "item_count",
    ]

    naturalness_related = [
        "naturalness",
    ]

    accuracy_related = [
        "accuracy",
    ]

    def calc_score(keys):
        failed_count = sum(1 for k in keys if checks[k][0] is False)

        if failed_count == 0:
            return 5
        if failed_count == 1:
            return 2
        return 1

    return {
        "auto_instruction_score": calc_score(instruction_related),
        "auto_completeness_score": calc_score(completeness_related),
        "auto_format_score": calc_score(format_related),
        "auto_naturalness_score": calc_score(naturalness_related),
        "auto_accuracy_score": calc_score(accuracy_related),
    }


def classify_error(checks: dict):
    failed_reasons = [
        reason for passed, reason in checks.values()
        if passed is False
    ]

    reason_text = "；".join(failed_reasons)

    if not failed_reasons:
        return "无明显错误"

    if "判断错误" in reason_text:
        return "判断错误"

    if "缺少必须包含关键词" in reason_text:
        return "指令违背"

    if "出现禁用词" in reason_text:
        return "指令违背"

    if "多余解释" in reason_text:
        return "指令违背"

    if "条数不符合" in reason_text:
        return "回答不完整"

    if "超过" in reason_text or "表格" in reason_text:
        return "格式/长度错误"

    if "夸张表达" in reason_text or "不够简单易懂" in reason_text:
        return "表达不符合要求"

    return "其他错误"

def evaluate_row(row):
    '''
    ---------------------------------------------
    每次拿 CSV 表格里的一行数据，自动完成一次完整评估。

    1. 读取这一行的 user_input 和 model_output
    2. 从 user_input 里解析规则
    3. 用规则检查 model_output
    4. 给这条回答打五项评分
    5. 判断合格/不合格
    6. 归因错误类型
    7. 输出判断理由
    -------------------------------------------
    '''
    user_input = str(row["user_input"])
    model_output = str(row["model_output"])

    rules = extract_rules(user_input)

    checks = {
        "item_count": check_item_count(model_output, rules),
        "length": check_length(model_output, rules),
        "required_keywords": check_required_keywords(model_output, rules),
        "forbidden_words": check_forbidden_words(model_output, rules),
        "no_explanation": check_no_explanation(model_output, rules),
        "table": check_table(model_output, rules),
        "naturalness": check_naturalness(user_input, model_output, rules),
        "accuracy": check_accuracy(user_input, model_output),
    }

    scores = score_from_checks(checks)

    total_score = round(sum(scores.values()) / len(scores), 2)

    failed_reasons = [
        reason for passed, reason in checks.values()
        if passed is False
    ]

    if failed_reasons:
        reason = "；".join(failed_reasons)
    else:
        reason = "；".join(reason for passed, reason in checks.values())

    error_type = classify_error(checks)

    if (
        total_score >= 4
        and scores["auto_instruction_score"] >= 4
        and scores["auto_format_score"] >= 4
    ):
        auto_result = "合格"
    else:
        auto_result = "不合格"

    return pd.Series({
        "parsed_rules": str(rules),
        **scores,
        "auto_total_score": total_score,
        "auto_result": auto_result,
        "auto_error_type": error_type,
        "auto_reason": reason,
    })